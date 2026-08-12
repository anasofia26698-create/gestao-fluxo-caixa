from openpyxl import load_workbook
from pathlib import Path

path = Path('/home/ubuntu/upload/FluxodecaixaDiárioORIGINAL(3).xlsx')
wb = load_workbook(path, read_only=True, data_only=True)
print('SHEETS:', wb.sheetnames)
for ws in wb.worksheets:
    print('SHEET:', ws.title)
    for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
        print(repr(row))
    print('DIMENSIONS:', ws.max_row, ws.max_column)
